import os
import openpyxl


class ExcelWriter:
    def __init__(self, excel_sheet_path=r'./Sheet/'):
        self.__sheet_paths = [excel_sheet_path + file for file in os.listdir(excel_sheet_path)]
        # Raise an error if no Excel sheet is found in the given path
        if len(self.__sheet_paths) == 0:
            raise FileNotFoundError('No Excel sheet found in the given path.')

    def sheet_number(self):
        return len(self.__sheet_paths)

    def get_keywords(self, i=0):
        keywords = []
        workbook = openpyxl.load_workbook(self.__sheet_paths[i])
        worksheet = workbook.active
        first_row = worksheet[1]
        for cell in first_row:
            keywords.append(cell.value)
        return list(filter(lambda x: x is not None, keywords))

    def write(self, results, i=0):
        workbook = openpyxl.load_workbook(self.__sheet_paths[i])
        worksheet = workbook.active

        # Iterate over the rows in the results list and write each cell value
        for row_idx, row in enumerate(results, start=1):
            for col_idx, value in enumerate(row, start=1):
                # Get the cell coordinates in A1 notation (e.g. "A1", "B2")
                cell = openpyxl.utils.get_column_letter(col_idx) + str(row_idx + 1)
                # Write the cell value to the worksheet
                worksheet[cell] = value

        # Save the changes to the workbook
        workbook.save(self.__sheet_paths[i])
